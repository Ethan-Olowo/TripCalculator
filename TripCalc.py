import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import requests
import os
import openpyxl
from openpyxl import Workbook
from config import ORS_API_KEY

base_dr = os.path.dirname(__file__)
ACTIVITIES_FILE = os.path.join(base_dr, "Activities.txt")

def get_coordinates(address):
    url = "https://api.openrouteservice.org/geocode/search"
    params = {"api_key": ORS_API_KEY, "text": address}
    response = requests.get(url, params=params)
    data = response.json()
    if response.status_code != 200 or "features" not in data:
        raise Exception(f"Error geocoding address: {data.get('error', 'Unknown error')}")
    try:
        feature = data["features"][0]
        coordinates = feature["geometry"]["coordinates"]
        location_name = feature["properties"]["label"]
        return (coordinates[1], coordinates[0]), location_name  # Return (lat, lon) and name
    except (IndexError, KeyError):
        raise Exception("Invalid geocoding response structure")

def get_road_distance(start, end):
    url = "https://api.openrouteservice.org/v2/directions/driving-car"
    headers = {"Authorization": ORS_API_KEY}
    body = {"coordinates": [[start[1], start[0]], [end[1], end[0]]]}
    response = requests.post(url, headers=headers, json=body)
    data = response.json()
    if response.status_code != 200 or "routes" not in data:
        raise Exception(f"Error fetching route distance: {data.get('error', 'Unknown error')}")
    try:
        route = data["routes"][0]
        distance_meters = route["summary"]["distance"]
        return distance_meters / 1000  # Convert to kilometers
    except (IndexError, KeyError):
        raise Exception("Invalid response structure from API")

def calculate_fuel_needed(distance, fuel_efficiency):
    return distance / fuel_efficiency

def load_activities():
    #Load activities and their file paths from the activities file.
    activities = {}
    if os.path.exists(ACTIVITIES_FILE):
        with open(ACTIVITIES_FILE, "r") as file:
            for line in file:
                name, path = line.strip().split("=", 1)
                activities[name] = path
    return activities

def save_activity(activity_name, file_path):
    #Save a new activity to the activities file.
    with open(ACTIVITIES_FILE, "a") as file:
        file.write(f"{activity_name}={file_path}\n")

def save_trip_to_excel(activity_name, trip_data):
    #Save trip data to an Excel file for the given activity.
    activities = load_activities()
    if activity_name in activities:
        file_path = activities[activity_name]
    else:
        # Create a new Excel file for the activity
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not file_path:
            raise Exception("No file selected for the new activity.")
        save_activity(activity_name, file_path)

    # Open or create the Excel file
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
    else:
        workbook = Workbook()

    # Add a new sheet for the trip
    sheet_name = f"Trip {len(workbook.sheetnames) + 1}"
    sheet = workbook.create_sheet(title=sheet_name)

    # Write trip data to the sheet
    headers = ["Leg", "Description", "Distance (km)","Ltrs/Km", "Fuel Needed (liters)", "Fuel Price ", "Cost"]
    sheet.append(headers)
    for leg in trip_data["legs"]:
        sheet.append(leg)
    sheet.append([])
    sheet.append(["Total Distance (km)", trip_data["total_distance"]])
    sheet.append(["Total Fuel Needed (liters)", trip_data["total_fuel"]])
    sheet.append(["Total Trip Cost ", trip_data["total_cost"]])

    # Save the workbook
    workbook.save(file_path)

def calculate_trip():
    try:
        # Get user inputs
        activity_name = activity_var.get().strip()
        if not activity_name:
            raise Exception("Please provide an activity name.")
        origin = origin_entry.get()
        stops = [stop.strip() for stop in stops_text.get("1.0", tk.END).splitlines() if stop.strip()]
        fuel_efficiency = float(fuel_efficiency_entry.get())
        fuel_price = float(fuel_price_entry.get())
        round_trip = round_trip_var.get()

        # Fetch coordinates and exact names for all locations
        locations = [origin] + stops
        coordinates = []
        location_names = []
        for loc in locations:
            coord, name = get_coordinates(loc)
            coordinates.append(coord)
            location_names.append(name)

        # Calculate distances and fuel for each leg
        results = f"Trip saved under activity: {activity_name}\n\n"
        total_distance = 0
        trip_legs = []
        for i in range(len(coordinates) - 1):
            leg_distance = get_road_distance(coordinates[i], coordinates[i + 1])
            total_distance += leg_distance
            fuel_needed_leg = calculate_fuel_needed(leg_distance, fuel_efficiency)
            leg_cost = fuel_needed_leg * fuel_price
            trip_legs.append([f"Leg {i + 1}", f"{location_names[i]} to {location_names[i + 1]}", leg_distance, fuel_efficiency, fuel_needed_leg, fuel_price, leg_cost])

            results += (f"Leg {i + 1}: {location_names[i]} → {location_names[i + 1]}\n"
                        f"  Distance: {leg_distance:.2f} km\n"
                        f"  Fuel Needed: {fuel_needed_leg:.2f} liters\n\n")
            

        # Add round trip distance if enabled
        if round_trip:
            return_distance = get_road_distance(coordinates[-1], coordinates[0])
            total_distance += return_distance
            fuel_needed_leg = calculate_fuel_needed(return_distance, fuel_efficiency)
            leg_cost = fuel_needed_leg * fuel_price
            trip_legs.append(["Return Leg", f"{location_names[-1]} to {location_names[0]}", return_distance, fuel_efficiency, fuel_needed_leg, fuel_price, leg_cost])
            results += (f"Return Leg: {location_names[-1]} → {location_names[0]}\n"
                        f"  Distance: {return_distance:.2f} km\n"
                        f"  Fuel Needed: {fuel_needed_leg:.2f} liters\n\n")

        # Calculate total fuel and cost
        total_fuel_needed = calculate_fuel_needed(total_distance, fuel_efficiency)
        total_cost = total_fuel_needed * fuel_price

        # Save trip data to Excel
        trip_data = {
            "legs": trip_legs,
            "total_distance": total_distance,
            "total_fuel": total_fuel_needed,
            "total_cost": total_cost,
        }
        save_trip_to_excel(activity_name, trip_data)

        # Display results
        results += (f"Total Distance: {total_distance:.2f} km\n"
                   f"Total Fuel Needed: {total_fuel_needed:.2f} liters\n"
                   f"Total Trip Cost: {total_cost:.2f} ")
        results_text.set(results)
    except Exception as e:
        messagebox.showerror("Error", str(e))

# GUI Setup
root = tk.Tk()
root.title("Trip Calculator")
root.geometry("600x700")

# Add icon to the app

root.iconbitmap(os.path.join(base_dr, "/AppIcon.icns"))

tk.Label(root, text="Activity:").pack(pady=5)
activities = load_activities()
activity_var = tk.StringVar()
activity_dropdown = ttk.Combobox(root, textvariable=activity_var, values=list(activities.keys()), width=50)
activity_dropdown.pack()

tk.Label(root, text="Origin:").pack(pady=5)
origin_entry = tk.Entry(root, width=50)
origin_entry.pack()

tk.Label(root, text="Stops (one per line):").pack(pady=5)
stops_text = tk.Text(root, height=5, width=50)
stops_text.pack()

tk.Label(root, text="Fuel Efficiency (km/l):").pack(pady=5)
fuel_efficiency_entry = tk.Entry(root, width=20)
fuel_efficiency_entry.pack()

tk.Label(root, text="Fuel Price :").pack(pady=5)
fuel_price_entry = tk.Entry(root, width=20)
fuel_price_entry.pack()

# Round Trip Checkbox
round_trip_var = tk.BooleanVar()
round_trip_checkbox = tk.Checkbutton(root, text="Round Trip", variable=round_trip_var)
round_trip_checkbox.pack(pady=5)

# Calculate Button
tk.Button(root, text="Calculate Trip", command=calculate_trip).pack(pady=10)

# Results Display
results_text = tk.StringVar()
results_label = tk.Label(root, textvariable=results_text, justify="left", wraplength=550)
results_label.pack(pady=10)

root.mainloop()
