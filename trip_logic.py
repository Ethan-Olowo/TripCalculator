import os

import openpyxl
import requests
from openpyxl import Workbook

try:
    from config import ORS_API_KEY as DEFAULT_ORS_API_KEY
except Exception:
    DEFAULT_ORS_API_KEY = ""


base_dir = os.path.dirname(__file__)
ACTIVITIES_FILE = os.path.join(base_dir, "Activities.txt")


def _resolve_api_key(api_key=None):
    key = (api_key or DEFAULT_ORS_API_KEY or "").strip()
    if not key:
        raise ValueError("OpenRouteService API key is missing. Add it in Settings.")
    return key


def _extract_error_message(payload, default_message):
    if isinstance(payload, dict):
        err = payload.get("error")
        if isinstance(err, dict):
            message = err.get("message")
            if message:
                return str(message)
        if isinstance(err, str) and err.strip():
            return err.strip()

        message = payload.get("message")
        if isinstance(message, str) and message.strip():
            return message.strip()

    return default_message


def get_coordinates(address, api_key=None):
    key = _resolve_api_key(api_key)
    url = "https://api.openrouteservice.org/geocode/search"
    params = {"api_key": key, "text": address}
    response = requests.get(url, params=params, timeout=30)
    data = response.json()

    if response.status_code != 200 or "features" not in data:
        msg = _extract_error_message(data, "Unable to geocode the address.")
        raise Exception(msg)

    try:
        feature = data["features"][0]
        coordinates = feature["geometry"]["coordinates"]
        location_name = feature["properties"]["label"]
        return (coordinates[1], coordinates[0]), location_name
    except (IndexError, KeyError):
        raise Exception("Invalid geocoding response structure")


def get_road_distance(start, end, api_key=None):
    key = _resolve_api_key(api_key)
    url = "https://api.openrouteservice.org/v2/directions/driving-car"
    headers = {"Authorization": key}
    body = {"coordinates": [[start[1], start[0]], [end[1], end[0]]]}
    response = requests.post(url, headers=headers, json=body, timeout=30)
    data = response.json()

    if response.status_code != 200 or "routes" not in data:
        msg = _extract_error_message(data, "Unable to fetch route distance.")
        raise Exception(msg)

    try:
        route = data["routes"][0]
        distance_meters = route["summary"]["distance"]
        return distance_meters / 1000
    except (IndexError, KeyError):
        raise Exception("Invalid response structure from API")


def calculate_fuel_needed(distance, fuel_efficiency):
    return distance / fuel_efficiency


def load_activities():
    activities = {}
    if os.path.exists(ACTIVITIES_FILE):
        with open(ACTIVITIES_FILE, "r", encoding="utf-8") as file:
            for line in file:
                line = line.strip()
                if not line or "=" not in line:
                    continue
                name, path = line.split("=", 1)
                activities[name] = path
    return activities


def save_activity(activity_name, file_path):
    with open(ACTIVITIES_FILE, "a", encoding="utf-8") as file:
        file.write(f"{activity_name}={file_path}\n")


def load_activity_trip_defaults(activity_name):
    activities = load_activities()
    if activity_name not in activities:
        raise ValueError("Selected activity is not registered.")

    file_path = activities[activity_name]
    if not os.path.exists(file_path):
        raise ValueError("The saved activity file could not be found.")

    workbook = openpyxl.load_workbook(file_path, data_only=True)
    if not workbook.sheetnames:
        raise ValueError("No trip data found in the activity workbook.")

    sheet = workbook[workbook.sheetnames[-1]]

    leg_rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        leg_label = row[0] if len(row) > 0 else None
        if leg_label is None:
            continue

        leg_label_str = str(leg_label).strip()
        if not leg_label_str:
            continue

        if leg_label_str.startswith("Total Distance"):
            break

        description = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if " to " not in description:
            continue

        fuel_eff = row[3] if len(row) > 3 else None
        fuel_price = row[5] if len(row) > 5 else None
        leg_rows.append((leg_label_str, description, fuel_eff, fuel_price))

    if not leg_rows:
        raise ValueError("No leg data was found in the selected activity file.")

    route_points = []
    for _, description, _, _ in leg_rows:
        start, end = description.split(" to ", 1)
        start = start.strip()
        end = end.strip()
        if not route_points:
            route_points.append(start)
        route_points.append(end)

    if len(route_points) < 2:
        raise ValueError("The activity file does not contain a complete route.")

    round_trip = False
    if len(route_points) > 2 and route_points[-1] == route_points[0]:
        round_trip = True
        route_points = route_points[:-1]

    origin = route_points[0]
    stops = route_points[1:]
    if not stops:
        raise ValueError("The activity file does not contain stops to load.")

    first_row = leg_rows[0]
    fuel_eff = first_row[2]
    fuel_price = first_row[3]

    try:
        fuel_eff = float(fuel_eff)
        fuel_price = float(fuel_price)
    except (TypeError, ValueError):
        raise ValueError("Fuel settings in the activity file are invalid.")

    return {
        "origin": origin,
        "stops": stops,
        "fuel_efficiency": fuel_eff,
        "fuel_price": fuel_price,
        "round_trip": round_trip,
        "file_path": file_path,
    }


def save_trip_to_excel(activity_name, trip_data, file_dialog_func=None, register_activity=True):
    if activity_name is None:
        if file_dialog_func is not None:
            file_path = file_dialog_func()
        else:
            raise Exception("No file dialog function provided for saving trip.")
        if not file_path:
            raise Exception("No file selected for saving trip.")
    else:
        activities = load_activities()
        if activity_name in activities:
            file_path = activities[activity_name]
        else:
            if file_dialog_func is not None:
                file_path = file_dialog_func()
            else:
                raise Exception("No file dialog function provided for new activity.")
            if not file_path:
                raise Exception("No file selected for the new activity.")
            if register_activity:
                save_activity(activity_name, file_path)

    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
    else:
        workbook = Workbook()

    sheet_name = f"Trip {len(workbook.sheetnames) + 1}"
    sheet = workbook.create_sheet(title=sheet_name)

    headers = ["Leg", "Description", "Distance (km)", "Ltrs/Km", "Fuel Needed (liters)", "Fuel Price", "Cost"]
    sheet.append(headers)
    for leg in trip_data["legs"]:
        sheet.append(leg)

    sheet.append([])
    sheet.append(["Total Distance (km)", trip_data["total_distance"]])
    sheet.append(["Total Fuel Needed (liters)", trip_data["total_fuel"]])
    sheet.append(["Total Trip Cost", trip_data["total_cost"]])

    workbook.save(file_path)


def process_trip(
    activity_name,
    origin,
    stops,
    fuel_efficiency,
    fuel_price,
    round_trip,
    file_dialog_func=None,
    save_to_excel=True,
    register_activity=True,
    api_key=None,
):
    if fuel_efficiency <= 0:
        raise ValueError("Fuel efficiency must be greater than 0.")

    if fuel_price < 0:
        raise ValueError("Fuel price cannot be negative.")

    if not origin:
        raise ValueError("Origin is required.")

    if not stops:
        raise ValueError("Add at least one stop.")

    if not register_activity:
        activity_name = None

    key = _resolve_api_key(api_key)

    locations = [origin] + stops
    coordinates = []
    location_names = []
    for loc in locations:
        coord, name = get_coordinates(loc, api_key=key)
        coordinates.append(coord)
        location_names.append(name)

    total_distance = 0
    trip_legs = []
    leg_summaries = []

    for i in range(len(coordinates) - 1):
        leg_distance = get_road_distance(coordinates[i], coordinates[i + 1], api_key=key)
        total_distance += leg_distance
        fuel_needed_leg = calculate_fuel_needed(leg_distance, fuel_efficiency)
        leg_cost = fuel_needed_leg * fuel_price
        description = f"{location_names[i]} to {location_names[i + 1]}"

        trip_legs.append([
            f"Leg {i + 1}",
            description,
            leg_distance,
            fuel_efficiency,
            fuel_needed_leg,
            fuel_price,
            leg_cost,
        ])

        leg_summaries.append(
            {
                "label": f"Leg {i + 1}",
                "from": location_names[i],
                "to": location_names[i + 1],
                "distance": leg_distance,
                "fuel": fuel_needed_leg,
                "cost": leg_cost,
            }
        )

    if round_trip:
        return_distance = get_road_distance(coordinates[-1], coordinates[0], api_key=key)
        total_distance += return_distance
        return_fuel = calculate_fuel_needed(return_distance, fuel_efficiency)
        return_cost = return_fuel * fuel_price
        description = f"{location_names[-1]} to {location_names[0]}"

        trip_legs.append([
            "Return Leg",
            description,
            return_distance,
            fuel_efficiency,
            return_fuel,
            fuel_price,
            return_cost,
        ])

        leg_summaries.append(
            {
                "label": "Return Leg",
                "from": location_names[-1],
                "to": location_names[0],
                "distance": return_distance,
                "fuel": return_fuel,
                "cost": return_cost,
            }
        )

    total_fuel_needed = calculate_fuel_needed(total_distance, fuel_efficiency)
    total_cost = total_fuel_needed * fuel_price

    trip_data = {
        "legs": trip_legs,
        "total_distance": total_distance,
        "total_fuel": total_fuel_needed,
        "total_cost": total_cost,
    }

    if save_to_excel:
        save_trip_to_excel(
            activity_name,
            trip_data,
            file_dialog_func=file_dialog_func,
            register_activity=register_activity,
        )

    return {
        "activity_name": activity_name,
        "origin": location_names[0],
        "destination": location_names[-1],
        "round_trip": round_trip,
        "legs": leg_summaries,
        "total_distance": total_distance,
        "total_fuel": total_fuel_needed,
        "total_cost": total_cost,
        "fuel_efficiency": fuel_efficiency,
        "fuel_price": fuel_price,
    }
